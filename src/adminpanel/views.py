from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from src.users.models import User
from django.views.generic import TemplateView
from django.db.models import Count, Q
from django.core.cache import cache

import io
import os
from django.conf import settings
from django.http import HttpResponse
from openpyxl import load_workbook


from src.houses.models import Apartment, House, Call_Master
from src.finance.models import Payment_Account, CashRegister
from src.users.models import User
# Create your views here.


class ListDashboard(TemplateView):
    template_name = 'dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = cache.get('dashboard_stats')

        if not ctx:
            ctx = {
                'apartment': Apartment.objects.count(),
                'house': House.objects.count(),
                'payment_accounts': Payment_Account.objects.filter(status='active').count(),
                'all_users': User.objects.exclude(user_status='inactive').count()
            }

        call_status = Call_Master.objects.aggregate(
            new=Count('id', filter=Q(status='new')),
            work=Count('id', filter=Q(status='in_progress'))
        )

        ctx['new_call_master'] = call_status['new']

        ctx['work_call_master'] = call_status['work']

        amounts = CashRegister.objects.values_list('sum', flat=True)

        account_balance = Payment_Account.objects.filter(status='active').values_list('balance', flat=True)

        ctx['all_money'] = sum(amounts) or "0.00"

        ctx['account_balance'] = sum(account_balance) or "0.00"

        cache.set('dashboard_stats', ctx, 300)

        return ctx

def genetate_receipt_view(request, pk):

    data = {
        'iban': 'UA123456789000000000000000000',
        'account_num': '8888888',
        'doc_number': '00000',
        'date': '25.08.2006',
        'payer_info': 'Aboba Abobovuch',
        'total_accrued': 14.88,
        'to_pay': 14.88,
        'services': [
            {'name': 'Горячая вода', 'tariff': 2.5, 'measurement': 'м3', 'indicator': 30, 'sum': 75}
        ]
    }

    template_path = os.path.join(settings.BASE_DIR, 'src', 'static', 'dist', 'excel', 'Пример_Квитанции.xlsx')
    wb = load_workbook(template_path)
    ws = wb.active

    ws['B1'] = data['iban']
    ws['H2'] = data['account_num']
    ws['J2'] = data['doc_number']
    ws['J3'] = data['date']
    ws['B5'] = data['payer_info']
    ws['B6'] = data['total_accrued']
    ws['B8'] = data['to_pay']

    ws['B10'] = data['iban']
    ws['H11'] = data['account_num']
    ws['J11'] = data['doc_number']
    ws['J12'] = data['date']
    ws['B14'] = data['payer_info']
    ws['B15'] = data['total_accrued']
    ws['B17'] = data['to_pay']

    for i, service in enumerate(data['services']):
        row = 19 + i
        ws[f'A{row}'] = service['name']
        ws[f'C{row}'] = f"Основний тариф {service['tariff']}"
        ws[f'E{row}'] = service['measurement']
        ws[f'G{row}'] = service['indicator']
        ws[f'I{row}'] = service['sum']

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="receipt_{data["doc_number"]}.xlsx"'

    return response